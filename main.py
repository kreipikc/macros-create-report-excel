import math
import os
import sys
import win32com.client
from datetime import timedelta
from typing import Optional, List
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, numbers, Font, Alignment
from pydantic import ValidationError
from babel.dates import format_date
from schemas import ChecksDefault, AdditionalInfo, TypeCheck
from utils import (
    create_check,
    sum_money_all_checks,
    create_text_price, validate_check,
    get_absolute_path, create_representative_word,
    convert_num_to_word, create_kopecks_str
)
from config import (
    POST_CELL,
    REPORT_MONTH_CELL,
    EMPLOYEE_CELL,
    START_ROW_READ,
    START_ROW_WRITE,
    COUNT_ROW_AFTER_CHECKS,
    LOCATE_DATE, DATE_REPORT, DEPARTMENT_CELL
)


def read_input_additional_info(file_path: str) -> Optional[AdditionalInfo]:
    """Read additional information from an Excel file.

    Args:
        file_path (str): The path to the Excel file containing additional information.

    Returns:
        Optional[AdditionalInfo]: An AdditionalInfo object containing employee, report month, and post data.

    Raises:
        ValidationError: If the data in the Excel file is invalid.
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        data = AdditionalInfo(
            employee=sheet[EMPLOYEE_CELL].value,
            date_report=sheet[DATE_REPORT].value,
            report_month=sheet[REPORT_MONTH_CELL].value,
            post=sheet[POST_CELL].value,
            department=sheet[DEPARTMENT_CELL].value
        )

        return data
    except ValidationError as e:
        print(f"Validation error: {e}")
        return None


def read_input_checks(file_path: str) -> List[ChecksDefault]:
    """Read check data from an Excel file and validate it.

    Args:
        file_path (str): The path to the Excel file containing check data.

    Returns:
        List[ChecksDefault]: A list of ChecksDefault objects created from the Excel data.

    Raises:
        Exception: If any check data is invalid or missing required fields.
    """
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=START_ROW_READ, values_only=True):
        if all(cell is None for cell in row):
            break

        check = create_check(row)
        if check:
            try:
                validate_check(check)
                data.append(check)
            except Exception as e:
                print(f"Ошибка валидации! {e}")

    # for row in data:
    #     print(row)

    return data


def create_report(checks: List[ChecksDefault], info_data: AdditionalInfo, path_save: str) -> None:
    """Create a report by filling a template with check data and additional information.

    Args:
        checks (List[ChecksDefault]): A list of ChecksDefault objects to be included in the report.
        info_data (AdditionalInfo): Additional information to be included in the report.
        path_save (str): The path to save the report.

    Returns:
        None

    Raises:
        Exception: If there is an error while creating the report.
    """
    workbook = load_workbook(get_absolute_path("templates\\template_advance_report.xlsx"))
    sheet = workbook.active
    sys.stdout.reconfigure(encoding='utf-8')

    sum_checks = sum_money_all_checks(checks)
    rubles = int(sum_checks)
    kopecks = int((sum_checks - rubles) * 100)

    sheet['R9'] = rubles
    sheet['X9'] = kopecks
    sheet['J13'] = info_data.date_report.strftime('%d.%m.%Y')
    sheet['O15'] = format_date(info_data.date_report, format='d MMMM yyyy г.', locale=LOCATE_DATE)
    sheet['H19'] = info_data.department
    sheet['F21'] = info_data.employee
    sheet['Q23'] = f"Расходы {format_date(info_data.report_month, format='MMMM yyyy', locale=LOCATE_DATE)}"
    sheet['J33'] = sum_checks
    sheet['J39'] = create_text_price(rubles, kopecks)
    sheet['I55'] = info_data.employee
    sheet['D56'] = format_date(info_data.date_report, format='d MMMM yyyy г.', locale=LOCATE_DATE)
    sheet['K56'] = create_text_price(rubles, kopecks)

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    for idx, check in enumerate(checks, start=START_ROW_WRITE):
        sheet.insert_rows(idx)
        sheet.row_dimensions[idx].height = 23

        sheet.merge_cells(f'B{idx}:C{idx}')
        sheet[f'B{idx}'].number_format = numbers.FORMAT_NUMBER
        sheet[f'B{idx}'] = check.number_str

        sheet.merge_cells(f'D{idx}:E{idx}')
        sheet[f'D{idx}'] = check.date.strftime('%d.%m.%Y') if check.date is not None else None
        sheet[f'D{idx}'].alignment = Alignment(vertical='top', horizontal='center')

        sheet.merge_cells(f'F{idx}:G{idx}')
        sheet[f'F{idx}'] = check.id_check if check.id_check is not None else None
        sheet[f'F{idx}'].alignment = Alignment(vertical='top', horizontal='left')

        sheet.merge_cells(f'H{idx}:K{idx}')
        sheet[f'H{idx}'] = check.type_document.value
        sheet[f'H{idx}'].alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

        sheet.merge_cells(f'L{idx}:N{idx}')
        sheet[f'L{idx}'].number_format = numbers.FORMAT_NUMBER
        sheet[f'L{idx}'] = check.sum_check
        sheet[f'L{idx}'].alignment = Alignment(vertical='top', horizontal='right')

        sheet.merge_cells(f'O{idx}:Q{idx}')

        sheet.merge_cells(f'R{idx}:T{idx}')
        sheet[f'R{idx}'].number_format = numbers.FORMAT_NUMBER
        sheet[f'R{idx}'] = check.sum_check
        sheet[f'R{idx}'].alignment = Alignment(vertical='top', horizontal='right')

        sheet.merge_cells(f'U{idx}:W{idx}')

        sheet.merge_cells(f'X{idx}:Y{idx}')

        for row in sheet[f'B{idx}:Y{idx}']:
            for cell in row:
                cell.border = border


    # Заполнение "Итого" и данных на этой строке
    new_block_data_row = START_ROW_WRITE + len(checks)

    for i in range(COUNT_ROW_AFTER_CHECKS):
        sheet.row_dimensions[new_block_data_row + i].height = 11

    sheet.merge_cells(f'H{new_block_data_row}:K{new_block_data_row}')
    sheet[f'H{new_block_data_row}'] = "Итого"

    sheet.merge_cells(f'L{new_block_data_row}:N{new_block_data_row}')
    sheet[f'L{new_block_data_row}'] = f"=SUM(L{START_ROW_WRITE}:L{new_block_data_row - 1})"

    sheet.merge_cells(f'O{new_block_data_row}:Q{new_block_data_row}')

    sheet.merge_cells(f'R{new_block_data_row}:T{new_block_data_row}')
    sheet[f'R{new_block_data_row}'] = f"=SUM(R{START_ROW_WRITE}:R{new_block_data_row - 1})"

    sheet.merge_cells(f'U{new_block_data_row}:W{new_block_data_row}')

    sheet.merge_cells(f'X{new_block_data_row}:Y{new_block_data_row}')

    for row in sheet[f'L{new_block_data_row}:W{new_block_data_row}']:
        for cell in row:
            cell.border = border


    # Создание мест для подписей
    sheet.merge_cells(f'B{new_block_data_row + 2}:E{new_block_data_row + 2}')
    sheet[f'B{new_block_data_row + 2}'] = "Подотчетное лицо"

    sheet.merge_cells(f'F{new_block_data_row + 2}:L{new_block_data_row + 2}')
    for row in sheet[f'F{new_block_data_row + 2}:L{new_block_data_row + 2}']:
        for cell in row:
            cell.border = Border(bottom=Side(border_style='thin', color='000000'))

    sheet.merge_cells(f'F{new_block_data_row + 3}:L{new_block_data_row + 3}')
    sheet[f'F{new_block_data_row + 3}'] = "подпись"
    cell_signature = sheet[f'F{new_block_data_row + 3}']
    cell_signature.font = Font(name='Arial', size=6, bold=False, color='000000')
    cell_signature.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells(f'N{new_block_data_row + 2}:Y{new_block_data_row + 2}')
    for row in sheet[f'N{new_block_data_row + 2}:Y{new_block_data_row + 2}']:
        for cell in row:
            cell.border = Border(bottom=Side(border_style='thin', color='000000'))

    sheet.merge_cells(f'N{new_block_data_row + 3}:Y{new_block_data_row + 3}')
    sheet[f'N{new_block_data_row + 3}'] = "расшифровка подписи"
    cell_decrypt_signature = sheet[f'N{new_block_data_row + 3}']
    cell_decrypt_signature.font = Font(name='Arial', size=6, bold=False, color='000000')
    cell_decrypt_signature.alignment = Alignment(horizontal='center', vertical='center')

    sheet.merge_cells(f'B{new_block_data_row + 6}:E{new_block_data_row + 6}')
    sheet[f'B{new_block_data_row + 6}'] = "Руководитель"

    sheet.merge_cells(f'F{new_block_data_row + 6}:L{new_block_data_row + 6}')
    for row in sheet[f'F{new_block_data_row + 6}:L{new_block_data_row + 6}']:
        for cell in row:
            cell.border = Border(bottom=Side(border_style='thin', color='000000'))

    sheet.merge_cells(f'N{new_block_data_row + 6}:Y{new_block_data_row + 6}')
    for row in sheet[f'N{new_block_data_row + 6}:Y{new_block_data_row + 6}']:
        for cell in row:
            cell.border = Border(bottom=Side(border_style='thin', color='000000'))

    sheet[f'N{new_block_data_row + 2}'] = info_data.employee
    sheet[f'N{new_block_data_row + 6}'] = info_data.employee

    workbook.save(f"{path_save}\\Авансовый отчет {info_data.date_report.strftime('%d-%m-%Y')}.xlsx")

    # EXCEL -> PDF
    excel = win32com.client.Dispatch("Excel.Application")

    workbook = excel.Workbooks.Open(get_absolute_path(f"{path_save}\\Авансовый отчет {info_data.date_report.strftime('%d-%m-%Y')}.xlsx"))

    worksheet = workbook.ActiveSheet
    worksheet.ExportAsFixedFormat(0, get_absolute_path(f"{path_save}\\Авансовый отчет {info_data.date_report.strftime('%d-%m-%Y')}.pdf"))

    workbook.Close(False)
    excel.Quit()


def create_additional_reports(checks: List[ChecksDefault], info_data: AdditionalInfo, path_save: str) -> None:
    """Function to generate additional reports based on check data and additional information.

    Args:
        checks (List[ChecksDefault]): A list of check objects containing data for report generation.
        info_data (AdditionalInfo): An object containing additional information required for the reports.
        path_save (str): The directory path where the generated reports will be saved.

    Returns:
        None
    """
    for check in checks:
        if check.type == TypeCheck.representative_offices_event:
            money = math.ceil(check.sum_check / 1000) * 1000
            replacements = {
                "{{counterparty}}": check.counterparty,
                "{{date_compilation}}": str(format_date((check.date - timedelta(days=7)), format='dd MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{date}}": str(format_date(check.date, format='dd MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{meeting_place}}": check.meeting_place,
                "{{post}}": info_data.post,
                "{{employee}}": info_data.employee,
                "{{counterparty_participant}}": check.counterparty_participant,
                "{{counterparty_post}}": check.counterparty_post,
                "{{budget}}": f"{money:.2f} рублей ({convert_num_to_word(int(money))} рублей {create_kopecks_str(check.sum_check)} копеек)",
                "{{day}}": str(format_date((check.date - timedelta(days=7)), format='dd', locale=LOCATE_DATE)),
                "{{month}}": str(format_date((check.date - timedelta(days=7)), format='MMMM', locale=LOCATE_DATE)),
                "{{year}}": str(format_date((check.date - timedelta(days=7)), format='yyyy', locale=LOCATE_DATE)),
                "{{price_num}}": f"{check.sum_check:.2f}",
                "{{price_str}}": f"{convert_num_to_word(int(check.sum_check))} рублей {create_kopecks_str(check.sum_check)} копеек",
                "{{date_compilation_2}}": str(format_date((check.date - timedelta(days=7)), format='«dd» MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{date_default}}": str(check.date.strftime('%d.%m.%Y')),
                "{{id}}": str(check.id_check),
            }
            print(f"Создание отчета 'Представительские_{check.id_check}'...")
            create_representative_word(replacements=replacements, file_template="template_representative.docx", output_path=f"{path_save}\\Представительские_{check.id_check}")
            print(f"Отчет 'Представительские_{check.id_check}' создан!")
        elif check.type == TypeCheck.representative_offices_present:
            money = math.ceil(check.sum_check / 1000) * 1000
            replacements = {
                "{{topic}}": check.topic,
                "{{date_compilation}}": str(format_date((check.date - timedelta(days=7)), format='dd MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{date}}": str(format_date(check.date, format='«dd» MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{post}}": info_data.post,
                "{{employee}}": info_data.employee,
                "{{counterparty}}": check.counterparty,
                "{{counterparty_participant}}": str(check.counterparty_participant),
                "{{budget}}": f"{money:.2f} рублей ({convert_num_to_word(int(money))} рублей {create_kopecks_str(check.sum_check)} копеек)",
                "{{day}}": str(format_date(check.date, format='dd', locale=LOCATE_DATE)),
                "{{month}}": str(format_date(check.date, format='MMMM', locale=LOCATE_DATE)),
                "{{year}}": str(format_date(check.date, format='yyyy', locale=LOCATE_DATE)),
                "{{name_present}}": check.name_present,
                "{{count_present}}": str(len([word.strip() for word in check.name_present.split(", ")])),
                "{{price}}": str(check.sum_check)
            }
            print(f"Создание отчета 'Представительские Подарки_{check.id_check}'...")
            create_representative_word(replacements=replacements, file_template="template_presents.docx", output_path=f"{path_save}\\Представительские Подарки_{check.id_check}")
            print(f"Отчет 'Представительские Подарки_{check.id_check}' создан!")
        elif check.type == TypeCheck.round_table_discussion_Club:
            money = math.ceil(check.sum_check / 1000) * 1000
            replacements = {
                "{{medication}}": check.medication,
                "{{date_compilation}}": str(format_date((check.date - timedelta(days=7)), format='dd MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{date}}": str(format_date(check.date, format='dd MMMM yyyy г.', locale=LOCATE_DATE)),
                "{{meeting_place}}": check.meeting_place,
                "{{post}}": info_data.post,
                "{{employee}}": info_data.employee,
                "{{topic}}": check.topic,
                "{{counterparty_participant}}": check.counterparty_participant,
                "{{counterparty_post}}": check.counterparty_post,
                "{{budget}}": f"{money:.2f} рублей ({convert_num_to_word(int(money))} рублей {create_kopecks_str(check.sum_check)} копеек)",
            }
            print(f"Создание отчета 'БЗ Круглый стол_{check.id_check}'...")
            create_representative_word(replacements=replacements, file_template="template_round_table.docx", output_path=f"{path_save}\\БЗ Круглый стол_{check.id_check}")
            print(f"Отчет 'БЗ Круглый стол_{check.id_check}' создан!")


def main(path_input_file: str, path_save: str) -> None:
    """Main function to process an Excel file and generate a report.

    Args:
        path_input_file (str): The file path to the Excel file containing check data and additional information.
        path_save (str): The path directory to save the report.

    Returns:
        None
    """
    print("Старт сканирования данных...")
    checks_all = read_input_checks(path_input_file)
    info = read_input_additional_info(path_input_file)
    print("Сканирование завершено!")
    print("Старт создание отчетов...")
    create_report(checks_all, info, path_save)
    create_additional_reports(checks_all, info, path_save)
    print("Создание отчетов завершено!")
    print("Можете закрывать консоль.")


if __name__ == "__main__":
    print("Запуск скрипта...")
    if len(sys.argv) > 2:
        file_path_input = sys.argv[1]
        file_path_save = sys.argv[2]

        if not os.path.exists(file_path_save):
            os.makedirs(file_path_save)

        main(file_path_input, file_path_save)
    else:
        print("Ошибка. Не переданы пути для работы скрипта.")